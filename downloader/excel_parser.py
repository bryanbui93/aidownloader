import re
from typing import List

import openpyxl

from .models import DownloadJob

# Regex to identify a cell as a video URL from supported platforms
_URL_PATTERN = re.compile(
    r"https?://(www\.)?"
    r"(tiktok\.com|douyin\.com|facebook\.com|fb\.watch|youtube\.com/shorts|youtu\.be)",
    re.IGNORECASE,
)


def _is_video_url(value: str) -> bool:
    return bool(_URL_PATTERN.search(str(value)))


def _detect_url_column(sheet) -> int:
    """
    Scan first 20 rows × all columns to find the column index
    with the highest density of video URLs. Returns 1-based column index.
    Raises ValueError if none found.
    """
    max_rows_to_scan = min(sheet.max_row, 20)
    scores: dict[int, int] = {}

    for row in sheet.iter_rows(min_row=1, max_row=max_rows_to_scan):
        for cell in row:
            if cell.value and _is_video_url(str(cell.value)):
                scores[cell.column] = scores.get(cell.column, 0) + 1

    if not scores:
        raise ValueError(
            "No supported video URLs found in the Excel file.\n"
            "Supported platforms: TikTok, Douyin, Facebook Reels, YouTube Shorts."
        )

    best_col = max(scores, key=lambda c: scores[c])
    return best_col


def parse_excel(filepath: str) -> List[DownloadJob]:
    """
    Parse an Excel file, auto-detect the URL column, and return
    a list of DownloadJob objects for all non-empty URL cells.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet = wb.active

    url_col = _detect_url_column(sheet)

    jobs: List[DownloadJob] = []
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1
    ):
        cell = row[url_col - 1]
        if cell.value and _is_video_url(str(cell.value)):
            url = str(cell.value).strip()
            jobs.append(DownloadJob(url=url, row_number=row_idx))

    wb.close()

    if not jobs:
        raise ValueError("URL column detected but no valid URLs were found.")

    return jobs
